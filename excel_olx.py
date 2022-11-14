import os
import shutil
from datetime import datetime
from log_print import Log
from olx_lib import Offer
from openpyxl import *
from openpyxl.styles import PatternFill, Color, Alignment

LENGTH_LIMIT = 500
MAX_COLUMNS = 15

class Offersheet:
    def __init__(self, filename: str):
        self.filename = filename
        self.titles = ['Name', 'Price', 'Location', 'Date', 'Gear', 'State', 'Notes', 'Link']
        self.columns_width = {
            'Name': 40,
            'Price': 20,
            'Location': 20,
            'Date': 20,
            'Link': 30,
            'Gear': 10,
            'State': 10,
            'Notes': 40,
        }
        if self.filename in os.listdir(os.path.abspath(os.curdir)):
            self.window = load_workbook(filename=self.filename)
            self.sheet = self.window.active
            try:
                shutil.copy2(filename, f"backup/{filename.replace('.xlsx', '')}_backup.xlsx")
            except IOError:
                Log("backup problem")
            for i in range(MAX_COLUMNS):
                column_name = self.sheet[f'{chr(ord("A") + i)}1'].value
                if column_name not in self.titles:
                    self.titles.append(column_name)

        else:
            self.window = self.create_sheet()
            self.sheet = self.window.active

    def create_sheet(self) -> Workbook:
        self.window = Workbook()
        self.sheet = self.window.active
        greybg = PatternFill(patternType='solid', fgColor=Color(rgb='808080'))
        for row in range(len(self.titles)):
            self.sheet[f'{chr(ord("A") + row)}1'] = self.titles[row]
            self.sheet[f'{chr(ord("A") + row)}1'].fill = greybg
            self.sheet[f'A3'] = "INACTIVE OFFERS"
            self.sheet[f'A3'].fill = greybg
            self.sheet.column_dimensions[chr(ord("A") + row)].width = self.columns_width[self.titles[row]]
        return self.window

    def find_first_empty_row(self, start_row: int = 1):
        for row in range(start_row, LENGTH_LIMIT):
            if self.sheet[f'A{row}'].value == '' or self.sheet[f'A{row}'].value is None:
                break
        return row

    def find_first_empty_row_in_inactive(self):
        startpoint = self.look_for_value(search_by_category="Name", value="INACTIVE OFFERS", whole_document=True)
        for row in range(startpoint, LENGTH_LIMIT):
            if self.sheet[f'A{row}'].value == '' or self.sheet[f'A{row}'].value is None:
                break
        return row

    def find_column(self, category: str):
        find_flag = False
        for i in range(len(self.titles)):
            if self.sheet[f'{chr(ord("A") + i)}1'].value == category:
                find_flag = True
                break
        if not find_flag:
            raise Exception
        return chr(ord("A") + i)

    def add_new_offer(self, offer: Offer, row: int = None, add_new_row=True):
        if row is None:
            row = self.find_first_empty_row()
        if add_new_row:
            self.sheet.insert_rows(idx=row, amount=1)
        self.sheet[f"{self.find_column('Name')}{str(row)}"] = offer.name
        self.sheet[f"{self.find_column('Price')}{str(row)}"] = offer.price
        self.sheet[f"{self.find_column('Location')}{str(row)}"] = offer.location
        self.sheet[f"{self.find_column('Date')}{str(row)}"] = offer.date
        self.sheet[f"{self.find_column('Link')}{str(row)}"] = offer.link

    def look_for_value(self, search_by_category: str, value, whole_document: bool = False) -> int:
        found_row = None
        column = self.find_column(search_by_category)

        for i in range(1, self.find_first_empty_row() if not whole_document else LENGTH_LIMIT):
            if self.sheet[f'{column}{i}'].value == value:
                found_row = i
                break
        return found_row

    def compare_offers(self, row: int, offer: Offer):
        changed = False
        ofert_excel_name = self.sheet[f"{self.find_column('Name')}{str(row)}"].value
        if ofert_excel_name != offer.name:
            changed = True
            self.sheet[
                f"{self.find_column('Name')}{str(row)}"] = f"{offer.name} - {datetime.now().strftime('%d/%m/%Y')} \n{ofert_excel_name}"
        ofert_excel_price = self.sheet[f"{self.find_column('Price')}{str(row)}"].value
        if ofert_excel_price != offer.price:
            changed = True
            self.sheet[
                f"{self.find_column('Price')}{str(row)}"] = f"{offer.price} - {datetime.now().strftime('%d/%m/%Y')} \n{ofert_excel_price}"
        ofert_excel_location = self.sheet[f"{self.find_column('Location')}{str(row)}"].value
        if ofert_excel_location != offer.location:
            changed = True
            self.sheet[
                f"{self.find_column('Location')}{str(row)}"] = f"{offer.location} - {datetime.now().strftime('%d/%m/%Y')} \n{ofert_excel_location}"
        ofert_excel_link = self.sheet[f"{self.find_column('Link')}{str(row)}"].value
        if ofert_excel_link != offer.link:
            changed = True
            self.sheet[
                f"{self.find_column('Link')}{str(row)}"] = f"{offer.link} - {datetime.now().strftime('%d/%m/%Y')} \n{ofert_excel_link}"
        if changed:
            self.highlight_row(row)
    def look_in_inactive(self, search_by_category: str, value) -> int:
        found_row = None
        column = self.find_column(search_by_category)
        startpoint = self.look_for_value(search_by_category="Name", value="INACTIVE OFFERS", whole_document=True)
        for i in range(startpoint, self.find_first_empty_row_in_inactive()):
            if self.sheet[f'{column}{i}'].value == value:
                found_row = i
                break
        return found_row

    def save_file(self):
        for row in range(1, LENGTH_LIMIT):
            for column in range(20):
                self.sheet[f"{chr(ord('A') + column)}{row}"].alignment = Alignment(wrap_text=True)
        self.window.save(self.filename)

    def move_row(self, frm: int, to: int):
        tmp = []
        for column in range(len(self.titles)):
            tmp.append(self.sheet[f"{chr(ord('A') + column)}{frm}"].value)

        self.sheet.insert_rows(idx=to, amount=1)
        for column in range(len(self.titles)):
            self.sheet[f"{chr(ord('A') + column)}{to}"] = tmp[column]
        self.sheet.delete_rows(idx=frm if frm < to else frm + 1, amount=1)
        self.highlight_row(to - 1 if frm < to else to)
        return to - 1 if frm < to else to

    def search_inactive_offers(self, offers: list[Offer]):
        links_from_excel = []
        links_from_olx = [offer.link for offer in offers]
        rows_to_remove = []
        for row in range(2, self.find_first_empty_row()):
            links_from_excel.append([row, self.sheet[f"{self.find_column('Link')}{str(row)}"].value])
        for link in links_from_excel:
            if link[1] not in links_from_olx:
                rows_to_remove.append(link[0])
        rows_to_remove.reverse()
        try:
            for row in rows_to_remove:
                self.move_row(frm=row, to=self.find_first_empty_row_in_inactive())
        except TypeError:
            pass

    def highlight_row(self, row):
        redbg = PatternFill(patternType='solid', fgColor=Color(rgb='eb4949'))
        for column in range(len(self.titles)):
            self.sheet[f'{chr(ord("A") + column)}{row}'].fill = redbg
# todo add macro to clear highlited offers

# todo think about sorting
